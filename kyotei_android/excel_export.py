"""
Excel出力モジュール（Android版）
PC版 kyotei_race_analysis.py のExcel生成ロジックを完全移植。
PC版とまったく同じ .xlsx を出力する。
matplotlib のフォントは同梱の NotoSansJP.ttf を使用する。
"""

import io
import os
import tempfile

# matplotlib の設定/キャッシュ書き込み先を確実に書き込み可能な場所へ
# （Android ではホーム配下に書けずクラッシュすることがあるため）
os.environ.setdefault("MPLCONFIGDIR", os.path.join(tempfile.gettempdir(), "mplconfig"))
try:
    os.makedirs(os.environ["MPLCONFIGDIR"], exist_ok=True)
except Exception:
    pass

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.patches as patches
from matplotlib import font_manager, rcParams

# 日本語フォント登録（Meiryo の代わりに同梱 NotoSansJP を使用）
_FONT_PATH = os.path.join(os.path.dirname(__file__), "NotoSansJP.ttf")
if os.path.exists(_FONT_PATH):
    try:
        font_manager.fontManager.addfont(_FONT_PATH)
        rcParams["font.family"] = font_manager.FontProperties(
            fname=_FONT_PATH).get_name()
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D


BOAT_BG = {
    1: "CCCCCC", 2: "222222", 3: "E83030",
    4: "1A7ACC", 5: "E8C000", 6: "30A040",
}
BOAT_FG = {
    1: "000000", 2: "FFFFFF", 3: "FFFFFF",
    4: "FFFFFF", 5: "000000", 6: "FFFFFF",
}
BOAT_MPL = {
    1: ("#CCCCCC", "#000000"),
    2: ("#222222", "#FFFFFF"),
    3: ("#E83030", "#FFFFFF"),
    4: ("#1A7ACC", "#FFFFFF"),
    5: ("#E8C000", "#000000"),
    6: ("#30A040", "#FFFFFF"),
}

_ST_MAX = 0.40


def _parse_st(val: str):
    v = str(val).strip()
    if not v or v in ("F", "S0", "S1", "K", "欠", "-"):
        return None
    v = v.lstrip("+")
    try:
        return float(v)
    except ValueError:
        return None


# ── グラフ生成 ────────────────────────────

def _is_dark(hex_color: str) -> bool:
    """背景色が暗色かどうか判定（テキスト色の選択に使用）"""
    c = hex_color.lstrip("#")
    if len(c) != 6:
        return False
    r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) < 128


def _lighten(hex_color: str, factor: float = 0.35) -> str:
    c = hex_color.lstrip("#")
    if len(c) != 6:
        return hex_color
    r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    return "#{:02X}{:02X}{:02X}".format(
        min(255, int(r + (255 - r) * factor)),
        min(255, int(g + (255 - g) * factor)),
        min(255, int(b + (255 - b) * factor)),
    )


def _darken(hex_color: str, factor: float = 0.30) -> str:
    c = hex_color.lstrip("#")
    if len(c) != 6:
        return hex_color
    r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    return "#{:02X}{:02X}{:02X}".format(
        int(r * (1 - factor)), int(g * (1 - factor)), int(b * (1 - factor)),
    )


def _make_st_hist(ax, st_vals: list, title: str, color: str):
    valid = [v for v in st_vals if v is not None]
    invalid_count = len(st_vals) - len(valid)
    steps = round(_ST_MAX / 0.01)
    bin_counts = [0] * steps
    overflow = 0
    for v in valid:
        idx = round(v * 100)
        if 0 <= idx < steps:
            bin_counts[idx] += 1
        else:
            overflow += 1
    x_labels = [f".{i:02d}" if i % 5 == 0 else "" for i in range(steps)]
    all_counts = [invalid_count] + bin_counts
    if overflow:
        all_counts[-1] += overflow
    all_labels = ["F/S/欠"] + x_labels
    bars = ax.bar(range(len(all_counts)), all_counts,
                  color=color, edgecolor="white", linewidth=0.3, width=0.85)
    ax.set_xticks(range(len(all_labels)))
    ax.set_xticklabels(all_labels, fontsize=6, rotation=45, ha="right")
    ax.set_title(title, fontsize=9, pad=4)
    ax.set_ylabel("回数", fontsize=7)
    ax.tick_params(axis="y", labelsize=7)
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax.set_facecolor("#F8F9FA")
    n = len(st_vals)
    xlabel = f"STタイム  (n={n}"
    if valid:
        xlabel += f", 平均={sum(valid)/len(valid):.3f}"
    ax.set_xlabel(xlabel + ")", fontsize=7)
    for bar, cnt in zip(bars, all_counts):
        if cnt > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
                    str(cnt), ha="center", va="bottom", fontsize=5)


def _make_rank_bar(ax, subset: list, title: str, color: str):
    """
    ST順位分布の棒グラフ。
    各棒の内側に「勝率 XX% / 3連 XX%」を表示。
    subset: コース絞り込み済みのレース辞書リスト
    """
    # ST順位ごとに集計
    rank_data: dict[int, dict] = {r: {"n": 0, "win": 0, "top3": 0} for r in range(1, 7)}
    for rec in subset:
        rk = str(rec.get("スタート順位", "")).strip()
        if not (rk.isdigit() and 1 <= int(rk) <= 6):
            continue
        rk = int(rk)
        result = str(rec.get("結果", "")).strip()
        rank_data[rk]["n"] += 1
        if result == "1":
            rank_data[rk]["win"] += 1
        if result in ("1", "2", "3"):
            rank_data[rk]["top3"] += 1

    labels = list(range(1, 7))
    counts = [rank_data[r]["n"] for r in labels]
    bars = ax.bar([str(r) for r in labels], counts,
                  color=color, edgecolor="white", linewidth=0.5)
    ax.set_title(title, fontsize=9, pad=4)
    ax.set_xlabel(f"ST順位  (n={sum(counts)})", fontsize=7)
    ax.set_ylabel("回数", fontsize=7)
    ax.tick_params(labelsize=7)
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax.set_facecolor("#F8F9FA")

    max_count = max(counts) if any(c > 0 for c in counts) else 1

    for bar, rk in zip(bars, labels):
        n = rank_data[rk]["n"]
        if n == 0:
            continue

        bx = bar.get_x() + bar.get_width() / 2
        bh = bar.get_height()

        # カウント数（バー上部）
        ax.text(bx, bh + 0.05, str(n),
                ha="center", va="bottom", fontsize=7)

        # 勝率・3連対率（バー内部オーバーレイ）
        win_r  = rank_data[rk]["win"]  / n * 100
        top3_r = rank_data[rk]["top3"] / n * 100
        label  = f"勝率{win_r:.0f}%\n3連{top3_r:.0f}%"

        # バーが十分な高さなら内側に表示、低い場合は省略
        if bh >= max_count * 0.20:
            ax.text(bx, bh * 0.5, label,
                    ha="center", va="center",
                    fontsize=7.5, color="white" if _is_dark(color) else "#333333",
                    fontweight="bold", rotation=0,
                    linespacing=1.4, zorder=5)


def _fig_to_png_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    buf.seek(0)
    data = buf.read()
    plt.close(fig)
    return data


def make_comparison_chart(players_data: list[dict]) -> bytes:
    """
    6選手のSTタイム分布 + ST順位分布を2列×6行で並べたグラフを生成。
    players_data: [{course, name, races}, ...]  (course順 1-6)
    """
    fig, axes = plt.subplots(6, 2, figsize=(14, 24))
    fig.patch.set_facecolor("#FFFFFF")
    fig.suptitle("6艇 スタート分析グラフ（コース別）", fontsize=14, fontweight="bold",
                 color="#1F4E79", y=1.001)

    course_colors = {
        1: "#CCCCCC", 2: "#444444", 3: "#E83030",
        4: "#1A7ACC", 5: "#E8C000", 6: "#30A040",
    }

    for row_idx, pd in enumerate(players_data):
        course  = pd["course"]
        boat_no = int(pd.get("boat_no", course))
        name = pd["name"]
        races = pd["races"]
        color = course_colors.get(boat_no, "#4472C4")

        # コース絞り込み
        subset = [r for r in races if str(r.get("コース", "")).strip() == str(course)]
        st_all = [_parse_st(r.get("スタートタイム", "")) for r in subset]

        ax_st = axes[row_idx][0]
        ax_rk = axes[row_idx][1]

        label = f"{course}コース  {name}  (対象{len(subset)}R)"
        _make_st_hist(ax_st, st_all, f"{label} STタイム分布", color)
        _make_rank_bar(ax_rk, subset, f"{label} ST順位分布", color)

    fig.tight_layout(pad=1.5)
    return _fig_to_png_bytes(fig)


# ── スタート予想図 定数 ──
_SPEED  = 22.0  # m/s（0.1秒 = 2.2m）
_BOAT_L = 2.9   # 船体長 [m]
_BOAT_H = 1.6   # 船体高 [m]（視認用）
_LANE_H = 2.8   # レーン中心間距離 [m]

# 背景の表示サイズ（Excel上のピクセル数）
_BG_DISP_W = 1400
_BG_DISP_H = 560


def _disp_px_to_emu(px: float) -> int:
    """表示ピクセル（96dpi基準）を Excel EMU に変換"""
    return int(px / 96 * 914400)


def _course_y(c: int) -> float:
    return -(c - 1) * _LANE_H


def _collect_boat_info(players_data: list[dict]) -> list[dict]:
    result = []
    for pd in players_data:
        course = pd["course"]
        name   = pd["name"]
        races  = pd["races"]
        subset = [r for r in races if str(r.get("コース", "")).strip() == str(course)]
        st_vals = [_parse_st(r.get("スタートタイム", "")) for r in subset]
        valid   = [v for v in st_vals if v is not None]
        avg_st  = sum(valid) / len(valid) if valid else 0.99
        f_count = sum(1 for r in subset
                      if str(r.get("スタートタイム", "")).strip() in ("F", "S0", "S1"))
        f_rate  = f_count / len(subset) if subset else 0.0
        result.append({
            "course":   course,
            "boat_no":  int(pd.get("boat_no", course)),
            "name":     name,
            "avg_st":   avg_st,
            "f_rate":   f_rate,
            "n":        len(subset),
            "valid_n":  len(valid),
            "bow_x":    -avg_st * _SPEED,
            "stern_x":  -avg_st * _SPEED - _BOAT_L,
            "y_center": _course_y(course),
            "de_ashi":  int(pd.get("de_ashi", 3)),
            "nobiashi": int(pd.get("nobiashi", 3)),
        })
    return result


def _formation_axes_bounds(boat_info: list[dict]):
    """描画データ範囲 (x_min, x_max, y_min, y_max) を実座標から算出"""
    MARGIN_LEFT  = 1.0
    MARGIN_RIGHT = 2.0
    min_stern = min(info["stern_x"] for info in boat_info)
    max_bow   = max(info["bow_x"]   for info in boat_info)
    x_min = min_stern - MARGIN_LEFT
    x_max = max(max_bow + MARGIN_RIGHT, MARGIN_RIGHT)
    y_max =  _LANE_H * 0.6
    y_min = -5 * _LANE_H - _LANE_H * 0.6
    return x_min, x_max, y_min, y_max


_NOBIASHI_STEP = 0.3  # 評価1段階あたりの補正距離 [m]


def _apply_boat_corrections(boat_info: list[dict],
                             manual_offsets: dict | None = None) -> list[dict]:
    """
    伸足評価による位置補正とユーザー手動オフセットを適用した新しい boat_info を返す。
    伸足評価1=補正なし、2=+0.3m、3=+0.6m、4=+0.9m、5=+1.2m（スタートライン方向）。
    """
    corrected = []
    for info in boat_info:
        c = dict(info)
        nobiashi = info.get("nobiashi", 1)
        auto_shift   = (nobiashi - 1) * _NOBIASHI_STEP
        manual_shift = (manual_offsets or {}).get(info["course"], 0.0)
        total = auto_shift + manual_shift
        c["bow_x"]   = info["bow_x"]   + total
        c["stern_x"] = info["stern_x"] + total
        corrected.append(c)
    return corrected


def _data_to_disp(dx: float, dy: float,
                   x_min: float, x_max: float,
                   y_min: float, y_max: float) -> tuple[float, float]:
    """データ座標 → 背景表示ピクセル座標（左上原点）"""
    px = (dx - x_min) / (x_max - x_min) * _BG_DISP_W
    py = (1.0 - (dy - y_min) / (y_max - y_min)) * _BG_DISP_H
    return px, py


def make_boat_performance_chart(players_data: list[dict]) -> bytes:
    """
    船足評価サマリー図。
    各艇の 出足/伸足 評価 (1-5) と平均STを横並び棒グラフで可視化。
    """
    RATING_CLR = {1: "#B22222", 2: "#D06000", 3: "#C8A800",
                  4: "#3E9B3E", 5: "#006400"}
    EMPTY_CLR  = "#E0E0E0"

    n = len(players_data)
    ROW_H   = 1.15
    TITLE_H = 1.0
    total_h = TITLE_H + ROW_H * (n + 1)   # title + header + rows

    FIG_W = 12.0
    FIG_H = total_h * 0.65
    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_xlim(0, 12)
    ax.set_ylim(0, total_h)
    ax.axis("off")

    # タイトル
    ax.text(6, total_h - 0.45, "船足評価サマリー",
            ha="center", va="center", fontsize=15, fontweight="bold",
            color="#1F4E79")

    # ── ヘッダー行 ──
    HDR_Y = total_h - TITLE_H - ROW_H * 0.5
    HDR_COLS = [
        (0.60, "艇番"),
        (2.20, "選手名"),
        (5.10, "出足評価 (1-5)"),
        (8.90, "伸足評価 (1-5)"),
        (11.4, "平均ST"),
    ]
    hdr_rect = patches.Rectangle((0, HDR_Y - ROW_H * 0.5), 12, ROW_H,
                                   facecolor="#1F4E79", edgecolor="none")
    ax.add_patch(hdr_rect)
    for hx, ht in HDR_COLS:
        ax.text(hx, HDR_Y, ht, ha="center", va="center",
                fontsize=9, fontweight="bold", color="white")

    def _draw_rating_pips(cx, cy, rating, seg_count=5, bar_len=3.2, bar_h=0.50):
        sw = bar_len / seg_count
        for i in range(seg_count):
            filled = i < rating
            fill = RATING_CLR.get(rating, "#888") if filled else EMPTY_CLR
            rect = patches.Rectangle(
                (cx - bar_len / 2 + i * sw + 0.03, cy - bar_h / 2),
                sw - 0.06, bar_h,
                facecolor=fill, edgecolor="white", linewidth=0.8)
            ax.add_patch(rect)
        ax.text(cx + bar_len / 2 + 0.3, cy, str(rating),
                ha="left", va="center", fontsize=11, fontweight="bold",
                color=RATING_CLR.get(rating, "#888"))

    # ── 各艇行 ──
    for i, pd in enumerate(players_data):
        course   = pd["course"]
        boat_no  = int(pd.get("boat_no", course))
        name     = pd["name"]
        races    = pd["races"]
        de_ashi  = int(pd.get("de_ashi", 3))
        nobiashi = int(pd.get("nobiashi", 3))

        subset  = [r for r in races if str(r.get("コース", "")).strip() == str(course)]
        st_vals = [_parse_st(r.get("スタートタイム", "")) for r in subset]
        valid   = [v for v in st_vals if v is not None]
        avg_st  = sum(valid) / len(valid) if valid else None

        bg_color, fg_color = BOAT_MPL.get(boat_no, ("#AAAAAA", "#000000"))

        row_y = HDR_Y - ROW_H * (i + 1)

        # 交互背景
        row_bg = "#F0F6FF" if i % 2 == 0 else "#FFFFFF"
        ax.add_patch(patches.Rectangle((0, row_y - ROW_H * 0.5), 12, ROW_H,
                                        facecolor=row_bg, edgecolor="#DDDDDD",
                                        linewidth=0.5))

        # 艇番セル（ボートカラー）
        ax.add_patch(patches.Rectangle((0.05, row_y - ROW_H * 0.44), 1.10, ROW_H * 0.88,
                                        facecolor=bg_color, edgecolor="none",
                                        linewidth=0))
        ax.text(0.60, row_y, str(boat_no),
                ha="center", va="center", fontsize=16, fontweight="bold",
                color=fg_color)

        # 選手名
        short = name[:5] if len(name) > 5 else name
        ax.text(2.20, row_y, short, ha="center", va="center",
                fontsize=10, color="#222222")

        # 出足評価バー
        _draw_rating_pips(5.10, row_y, de_ashi)

        # 伸足評価バー
        _draw_rating_pips(8.90, row_y, nobiashi)

        # 平均ST
        st_txt = f".{round(avg_st * 1000):03d}" if avg_st is not None else "―"
        ax.text(11.4, row_y, st_txt, ha="center", va="center",
                fontsize=11, color="#333333", fontweight="bold")

    # 外枠
    ax.add_patch(patches.Rectangle((0, HDR_Y - ROW_H * n - ROW_H * 0.5),
                                    12, ROW_H * (n + 1),
                                    facecolor="none", edgecolor="#9EACC0",
                                    linewidth=1.2))

    fig.tight_layout(pad=0.3)
    return _fig_to_png_bytes(fig)


def _make_formation_background(boat_info: list[dict],
                                show_start_line: bool = True) -> bytes:
    """
    水面背景 + レーン線 + スタートライン（任意）のみを描画。
    テキスト・矢印・ラベルは一切含まない。
    axes が figure 全体を占めるため、データ座標→ピクセル変換が線形になる。
    """
    x_min, x_max, y_min, y_max = _formation_axes_bounds(boat_info)

    # axes が figure 全体を占めるよう figsize を合わせる（96dpi固定）
    FIG_DPI = 96
    fig = plt.figure(figsize=(_BG_DISP_W / FIG_DPI, _BG_DISP_H / FIG_DPI),
                     dpi=FIG_DPI)
    ax = fig.add_axes([0, 0, 1, 1])   # axes = figure 全体（余白ゼロ）
    fig.patch.set_facecolor("#4BA8D0")
    ax.set_facecolor("#4BA8D0")
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(y_min, y_max)
    ax.axis("off")

    # レーン区切り線
    for c in range(1, 8):
        yb = _course_y(c) + _LANE_H / 2
        ax.axhline(yb, color="white", linewidth=1.0, alpha=0.5, zorder=1)

    # スタートライン（赤破線）
    if show_start_line:
        ax.plot([0, 0], [y_min, y_max],
                color="#FF2222", linewidth=2.5, linestyle="--",
                alpha=0.95, zorder=3)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=FIG_DPI)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def make_boat_image_single(boat_no: int, name: str, avg_st: float,
                            f_warn: bool, valid_n: int,
                            show_labels: bool,
                            boat_w_px: int, boat_h_px: int) -> bytes:
    """
    1艇分の透過PNG（上面視・詳細グラフィック）。
    船首=右、船尾=左。流線型船体＋カラーデッキ＋コックピット＋水飛沫。
    """
    bg_color, fg_color = BOAT_MPL.get(boat_no, ("#AAAAAA", "#000000"))
    BL, BH = _BOAT_L, _BOAT_H

    FIG_DPI = 96
    fig = plt.figure(figsize=(boat_w_px / FIG_DPI, boat_h_px / FIG_DPI), dpi=FIG_DPI)
    ax = fig.add_axes([0, 0, 1, 1])
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")
    ax.set_xlim(0, BL)
    ax.set_ylim(-BH / 2, BH / 2)
    ax.axis("off")

    # 正規化座標 (nx∈[0,1], ny∈[0,1]) → データ座標
    def p(nx, ny):
        return nx * BL, (ny - 0.5) * BH

    # ピクセル半径 r_px → データ空間での楕円 (width, height) ─ 視覚的円になる
    def ell(r_px):
        return 2 * r_px * BL / boat_w_px, 2 * r_px * BH / boat_h_px

    hull_color  = "#BF8A12"
    hull_dark   = "#7A5509"
    hull_hi     = "#E0B040"

    # ── 船尾 水飛沫 ──
    for dy, alpha in [(-0.21, 0.75), (0.21, 0.75), (-0.37, 0.50), (0.37, 0.50)]:
        cx, cy = p(0.06, 0.50 + dy)
        ew, eh = ell(max(2, boat_h_px * 0.08))
        ax.add_patch(patches.Ellipse((cx, cy), ew * 2.8, eh,
                                      facecolor="#B8E0FF", alpha=alpha,
                                      edgecolor="white", linewidth=0.3, zorder=1))

    # ── 船体外形（流線型・上面視）──
    # 正規化 x:0=船尾, 1=船首 / y:0.50=中心線
    hull_pts = [
        # 船尾→下縁→船首
        (0.00, 0.12), (0.06, 0.06), (0.18, 0.03), (0.35, 0.03),
        (0.52, 0.05), (0.66, 0.09), (0.79, 0.16), (0.90, 0.26),
        (0.97, 0.38), (1.00, 0.50),
        # 船首→上縁→船尾
        (0.97, 0.62), (0.90, 0.74), (0.79, 0.84), (0.66, 0.91),
        (0.52, 0.95), (0.35, 0.97), (0.18, 0.97), (0.06, 0.94),
        (0.00, 0.88),
    ]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in hull_pts], closed=True,
                                  facecolor=hull_color, edgecolor=hull_dark,
                                  linewidth=0.8, zorder=2))

    # 船体内側シェーディング（立体感）
    inner_pts = [
        (0.02, 0.18), (0.08, 0.12), (0.20, 0.09), (0.36, 0.09),
        (0.52, 0.11), (0.66, 0.15), (0.78, 0.22), (0.88, 0.31),
        (0.94, 0.42), (0.96, 0.50),
        (0.94, 0.58), (0.88, 0.69), (0.78, 0.78), (0.66, 0.85),
        (0.52, 0.89), (0.36, 0.91), (0.20, 0.91), (0.08, 0.88),
        (0.02, 0.82),
    ]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in inner_pts], closed=True,
                                  facecolor=_darken(hull_color, 0.12),
                                  edgecolor="none", linewidth=0, zorder=3))

    # 上縁ハイライト線（光沢）
    hi_pts = [(0.02, 0.82), (0.20, 0.91), (0.52, 0.89), (0.78, 0.78),
              (0.94, 0.58), (0.96, 0.50),
              (0.94, 0.42), (0.78, 0.22), (0.52, 0.11), (0.20, 0.09),
              (0.02, 0.18), (0.00, 0.12), (0.00, 0.88)]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in hi_pts], closed=False,
                                  facecolor="none", edgecolor=hull_hi,
                                  linewidth=0.6, alpha=0.6, zorder=3))

    # ── カラーデッキ（艇番色） ──
    deck_pts = [
        (0.04, 0.20), (0.04, 0.80),
        (0.18, 0.88), (0.40, 0.90), (0.58, 0.86), (0.72, 0.78),
        (0.84, 0.64), (0.88, 0.50),
        (0.84, 0.36), (0.72, 0.22), (0.58, 0.14), (0.40, 0.10),
        (0.18, 0.12),
    ]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in deck_pts], closed=True,
                                  facecolor=bg_color, edgecolor="none",
                                  linewidth=0, zorder=4))

    # デッキ上部ハイライト（光沢ストライプ）
    hl_pts = [
        (0.04, 0.56), (0.04, 0.78), (0.18, 0.86), (0.40, 0.87),
        (0.58, 0.83), (0.72, 0.75), (0.82, 0.62), (0.86, 0.50),
    ]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in hl_pts], closed=False,
                                  facecolor=_lighten(bg_color, 0.22),
                                  edgecolor=_lighten(bg_color, 0.22),
                                  linewidth=1.2, alpha=0.40, zorder=5))

    # ── コックピット（暗いドーム） ──
    ck_cx, ck_cy = p(0.60, 0.50)
    ck_w, ck_h   = ell(max(4, boat_h_px * 0.24))
    ax.add_patch(patches.Ellipse((ck_cx, ck_cy),
                                  ck_w * 2.0, ck_h * 0.95,
                                  facecolor="#111111", edgecolor="#3A3A3A",
                                  linewidth=0.6, zorder=6))
    # コックピット内ガラス反射
    gl_cx, gl_cy = p(0.63, 0.43)
    gl_w, gl_h   = ell(max(1, boat_h_px * 0.07))
    ax.add_patch(patches.Ellipse((gl_cx, gl_cy), gl_w, gl_h,
                                  facecolor="white", alpha=0.22,
                                  edgecolor="none", zorder=7))

    # ── ヘルメット ──
    hm_cx, hm_cy = p(0.60, 0.50)
    hm_w, hm_h   = ell(max(3, boat_h_px * 0.16))
    ax.add_patch(patches.Ellipse((hm_cx, hm_cy), hm_w, hm_h,
                                  facecolor="#1C1C1C", edgecolor="#555555",
                                  linewidth=0.4, zorder=8))
    # ヘルメットシールド光沢
    sh_cx, sh_cy = p(0.62, 0.45)
    sh_w, sh_h   = ell(max(1, boat_h_px * 0.055))
    ax.add_patch(patches.Ellipse((sh_cx, sh_cy), sh_w, sh_h,
                                  facecolor="white", alpha=0.30,
                                  edgecolor="none", zorder=9))

    # ── 船首 水飛沫（白いV字） ──
    bow_spray = [(0.96, 0.44), (0.99, 0.40), (1.00, 0.50)]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in bow_spray], closed=True,
                                  facecolor="white", alpha=0.80, zorder=9))
    bow_spray2 = [(0.96, 0.56), (0.99, 0.60), (1.00, 0.50)]
    ax.add_patch(patches.Polygon([p(x, y) for x, y in bow_spray2], closed=True,
                                  facecolor="white", alpha=0.80, zorder=9))

    # ── 艇番 ──
    num_x = 0.30 if not show_labels else 0.26
    num_fs = max(5, int(boat_h_px * 0.28))
    ax.text(*p(num_x, 0.50), str(boat_no),
            ha="center", va="center",
            fontsize=num_fs, fontweight="bold",
            color=fg_color, zorder=10)

    # ── ラベル（選手名・ST） ──
    if show_labels:
        sub_fs = max(4, int(boat_h_px * 0.13))
        short = name[:4] if len(name) > 4 else name
        ax.text(*p(0.74, 0.78), short,
                ha="center", va="center",
                fontsize=sub_fs, color=fg_color, zorder=10)
        st_lbl = f".{round(avg_st * 1000):03d}" if valid_n > 0 else "―"
        ax.text(*p(0.74, 0.22), st_lbl,
                ha="center", va="center",
                fontsize=sub_fs, color=fg_color, zorder=10)

    # ── F警告 ──
    if f_warn:
        ax.text(*p(0.88, 0.22), "F",
                ha="center", va="center",
                fontsize=max(4, int(boat_h_px * 0.12)),
                color="#FF0000", fontweight="bold",
                bbox=dict(boxstyle="circle,pad=0.04", facecolor="white",
                          edgecolor="#FF0000", linewidth=0.7),
                zorder=11)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=FIG_DPI, transparent=True)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _embed_formation_in_sheet(ws, boat_info: list[dict],
                               bg_png: bytes, show_labels: bool,
                               bg_top_emu: int):
    """
    背景 PNG と各艇 PNG を AbsoluteAnchor で ws に貼り付ける。
    bg_top_emu: 背景画像の上端 EMU（シート原点からの距離）。
    """
    BG_LEFT_EMU = 0
    BG_W_EMU = _disp_px_to_emu(_BG_DISP_W)
    BG_H_EMU = _disp_px_to_emu(_BG_DISP_H)

    # 背景配置
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(bg_png)
        bg_path = tmp.name
    bg_img = XLImage(bg_path)
    bg_img.width  = _BG_DISP_W
    bg_img.height = _BG_DISP_H
    bg_img.anchor = AbsoluteAnchor(
        pos=XDRPoint2D(x=BG_LEFT_EMU, y=bg_top_emu),
        ext=XDRPositiveSize2D(cx=BG_W_EMU, cy=BG_H_EMU)
    )
    ws.add_image(bg_img)

    # 座標変換パラメータ
    x_min, x_max, y_min, y_max = _formation_axes_bounds(boat_info)

    # 艇の表示サイズ: x方向は物理スケール通り、y方向はレーン幅に合わせる
    boat_disp_w = max(60, int(_BOAT_L / (x_max - x_min) * _BG_DISP_W))
    boat_disp_h = max(20, int(_BOAT_H / (y_max - y_min) * _BG_DISP_H))

    for info in boat_info:
        course   = info["course"]
        boat_no  = info.get("boat_no", course)
        name     = info["name"]
        avg_st   = info["avg_st"]
        f_rate   = info["f_rate"]
        valid_n  = info["valid_n"]
        n        = info["n"]
        bow_x    = info["bow_x"]
        stern_x  = info["stern_x"]
        y_center = info["y_center"]
        f_warn   = f_rate >= 0.05 and n >= 5

        boat_png = make_boat_image_single(
            boat_no, name, avg_st, f_warn, valid_n,
            show_labels=show_labels,
            boat_w_px=boat_disp_w, boat_h_px=boat_disp_h)

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(boat_png)
            boat_path = tmp.name

        boat_img = XLImage(boat_path)
        boat_img.width  = boat_disp_w
        boat_img.height = boat_disp_h

        # 艇の左上ピクセル座標: stern_x が画像左端, y_center+BOAT_H/2 が画像上端
        img_left_px, img_top_px = _data_to_disp(
            stern_x, y_center + _BOAT_H / 2,
            x_min, x_max, y_min, y_max)

        emu_x = BG_LEFT_EMU + _disp_px_to_emu(max(0.0, img_left_px))
        emu_y = bg_top_emu  + _disp_px_to_emu(max(0.0, img_top_px))

        boat_img.anchor = AbsoluteAnchor(
            pos=XDRPoint2D(x=emu_x, y=emu_y),
            ext=XDRPositiveSize2D(cx=_disp_px_to_emu(boat_disp_w),
                                  cy=_disp_px_to_emu(boat_disp_h))
        )
        ws.add_image(boat_img)


# ── スタンドアロン PNG エクスポート ────────────────────────────

def make_clean_formation_png(players_data: list[dict]) -> bytes:
    """
    スタート配置図をmatplotlibで直接描画してPNG bytesで返す。
    Excelを介さない独立版（note添付用）。
    """
    boat_info_raw = _collect_boat_info(players_data)
    boat_info = _apply_boat_corrections(boat_info_raw)

    x_min, x_max, y_min, y_max = _formation_axes_bounds(boat_info)

    fig, ax = plt.subplots(figsize=(14, 5), dpi=120)
    fig.patch.set_facecolor("#3A9DC0")
    ax.set_facecolor("#3A9DC0")
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(y_min, y_max)
    ax.axis("off")

    fig.suptitle("スタート配置図", fontsize=14, fontweight="bold",
                 color="white", y=0.99)

    # レーン区切り線
    for c in range(1, 8):
        yb = _course_y(c) + _LANE_H / 2
        ax.axhline(yb, color="white", linewidth=1.0, alpha=0.45, zorder=1)

    # スタートライン
    ax.plot([0, 0], [y_min, y_max],
            color="#FF3333", linewidth=2.5, linestyle="--", alpha=0.9, zorder=3)
    ax.text(0.12, y_max * 0.85, "START",
            color="#FF3333", fontsize=8, fontweight="bold", va="center", zorder=5)

    # 各艇
    for info in boat_info:
        course   = info["course"]
        boat_no  = info.get("boat_no", course)
        name     = info["name"]
        bow_x    = info["bow_x"]
        stern_x  = info["stern_x"]
        y_center = info["y_center"]
        avg_st   = info.get("avg_st", 0)
        f_rate   = info.get("f_rate", 0)

        bg_color, fg_color = BOAT_MPL.get(boat_no, ("#AAAAAA", "#000000"))

        # 船体（角丸矩形）
        rect = patches.FancyBboxPatch(
            (stern_x, y_center - _BOAT_H / 2),
            _BOAT_L, _BOAT_H,
            boxstyle="round,pad=0.06",
            facecolor=bg_color, edgecolor="white",
            linewidth=1.5, zorder=4,
        )
        ax.add_patch(rect)

        cx = (stern_x + bow_x) / 2

        # 艇番（大）
        ax.text(cx, y_center, str(boat_no),
                ha="center", va="center", fontsize=15, fontweight="bold",
                color=fg_color, zorder=6)

        # 選手名（艇上部）
        short = name[:4] if len(name) > 4 else name
        ax.text(cx, y_center + _BOAT_H / 2 + 0.12, short,
                ha="center", va="bottom", fontsize=7.5,
                color="white", fontweight="bold", zorder=5,
                bbox=dict(boxstyle="round,pad=0.1", facecolor="#00000055",
                          edgecolor="none"))

        # 平均ST（艇の右 = 船首側）
        if avg_st > 0:
            st_str = f"ST .{round(avg_st * 1000):03d}"
            f_str  = f"  F{f_rate:.0%}" if f_rate >= 0.05 else ""
            ax.text(bow_x + 0.15, y_center, st_str + f_str,
                    ha="left", va="center", fontsize=7,
                    color="#FFFFAA", fontweight="bold", zorder=5)

    fig.tight_layout(pad=0.5, rect=[0, 0, 1, 0.96])
    return _fig_to_png_bytes(fig)


def export_analysis_pngs(
    players_data: list[dict],
    save_dir: str,
    prefix: str = "",
) -> dict[str, str]:
    """
    3種類のPNGをsave_dirに保存し、ファイルパスのdictを返す。
    キー: "boat_speed" / "st_graph" / "formation"
    """
    os.makedirs(save_dir, exist_ok=True)
    pfx = f"{prefix}_" if prefix else ""

    paths: dict[str, str] = {}

    p_speed = os.path.join(save_dir, f"{pfx}船足評価.png")
    with open(p_speed, "wb") as f:
        f.write(make_boat_performance_chart(players_data))
    paths["boat_speed"] = p_speed

    p_graph = os.path.join(save_dir, f"{pfx}スタート分析グラフ.png")
    with open(p_graph, "wb") as f:
        f.write(make_comparison_chart(players_data))
    paths["st_graph"] = p_graph

    p_form = os.path.join(save_dir, f"{pfx}スタート配置図.png")
    with open(p_form, "wb") as f:
        f.write(make_clean_formation_png(players_data))
    paths["formation"] = p_form

    return paths


# ── Excel出力 ─────────────────────────────

def write_analysis_excel(players_data: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    # ── シート1: 選手別出走データ ──
    _write_data_sheet(wb, players_data)

    # ── シート2: 比較グラフ ──
    _write_chart_sheet(wb, players_data)

    # ── シート3: スタート予想図 ──
    _write_formation_sheet(wb, players_data)

    # ── シート4: スタート配置図（伸足補正適用・艇番のみ）──
    _write_clean_formation_sheet(wb, players_data)

    # ── シート5: 船足評価 ──
    _write_performance_sheet(wb, players_data)

    wb.save(output_path)


def _write_data_sheet(wb, players_data: list[dict]):
    ws = wb.active
    ws.title = "出走データ"
    ws.sheet_view.showGridLines = True

    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    row = 1
    ws.merge_cells(f"A{row}:I{row}")
    tc = ws[f"A{row}"]
    tc.value = "6艇 出走データ（コース別）"
    tc.font = Font(bold=True, size=13, color="1F4E79")
    tc.alignment = center
    ws.row_dimensions[row].height = 28
    row += 1

    for pd in players_data:
        course  = pd["course"]
        boat_no = int(pd.get("boat_no", course))
        name    = pd["name"]
        races   = pd["races"]
        subset  = [r for r in races if str(r.get("コース", "")).strip() == str(course)]

        if not subset:
            continue

        bg = BOAT_BG.get(boat_no, "AAAAAA")
        fg = BOAT_FG.get(boat_no, "000000")

        # コースヘッダー
        ws.merge_cells(f"A{row}:I{row}")
        hc = ws[f"A{row}"]
        hc.value = f"【{course}コース / {boat_no}号艇】  {name}  （対象レース数: {len(subset)}）"
        hc.font = Font(bold=True, size=11, color=fg)
        hc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        # カラムヘッダー
        cols = ["競艇場", "開催日時", "開催レース", "艇番", "コース",
                "スタートタイム", "スタート順位", "結果(着)"]
        widths = [12, 13, 11, 6, 6, 14, 12, 9]
        for ci, (cn, cw) in enumerate(zip(cols, widths), 1):
            cell = ws.cell(row=row, column=ci, value=cn)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = max(
                ws.column_dimensions[get_column_letter(ci)].width or 0, cw
            )
        ws.row_dimensions[row].height = 20
        row += 1

        zebra = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")
        result_fill_map = {"1": "FFD700", "2": "C0C0C0", "3": "CD7F32"}

        for i, race in enumerate(subset):
            vals = [
                race.get("競艇場", ""), race.get("開催日時", ""),
                race.get("開催レース", ""), race.get("艇番", ""),
                race.get("コース", ""), race.get("スタートタイム", ""),
                race.get("スタート順位", ""), race.get("結果", ""),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = center
                cell.border = border
                if i % 2 == 0:
                    cell.fill = zebra

            bn = str(race.get("艇番", "")).strip()
            if bn in [str(x) for x in range(1, 7)]:
                b = int(bn)
                bbg = BOAT_BG.get(b, "AAAAAA")
                bfg = BOAT_FG.get(b, "000000")
                for ci in [4, 5]:
                    bc = ws.cell(row=row, column=ci)
                    bc.fill = PatternFill(start_color=bbg, end_color=bbg, fill_type="solid")
                    bc.font = Font(bold=True, color=bfg)

            rs = str(race.get("結果", "")).strip()
            if rs in result_fill_map:
                rc = ws.cell(row=row, column=8)
                rc.fill = PatternFill(
                    start_color=result_fill_map[rs],
                    end_color=result_fill_map[rs], fill_type="solid"
                )
                rc.font = Font(bold=True)

            ws.row_dimensions[row].height = 17
            row += 1

        row += 1  # 艇間スペース


def _write_chart_sheet(wb, players_data: list[dict]):
    ws = wb.create_sheet("スタート分析グラフ")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    tc = ws["A1"]
    tc.value = "6艇 スタート分析グラフ（コース別）"
    tc.font = Font(bold=True, size=13, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    png_data = make_comparison_chart(players_data)

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(png_data)
        tmp_path = tmp.name
    img = XLImage(tmp_path)
    img.width = 1100
    img.height = 1900
    ws.add_image(img, "A2")

    ws.column_dimensions["A"].width = 150


def _write_formation_sheet(wb, players_data: list[dict]):
    ws = wb.create_sheet("スタート予想図")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    tc = ws["A1"]
    tc.value = "スタート予想図"
    tc.font = Font(bold=True, size=13, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # サマリーテーブル
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["コース", "選手名", "対象R数", "平均ST", "F率", "ST1位率"]
    widths = [8, 12, 10, 10, 8, 10]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    for i, pd in enumerate(players_data):
        course  = pd["course"]
        boat_no = int(pd.get("boat_no", course))
        name    = pd["name"]
        races   = pd["races"]
        subset  = [r for r in races if str(r.get("コース", "")).strip() == str(course)]
        st_vals = [_parse_st(r.get("スタートタイム", "")) for r in subset]
        valid = [v for v in st_vals if v is not None]
        avg_st = sum(valid) / len(valid) if valid else None
        f_count = sum(1 for r in subset
                      if str(r.get("スタートタイム", "")).strip() in ("F", "S0", "S1"))
        f_rate = f_count / len(subset) if subset else 0.0
        rank_vals = [r.get("スタート順位", "") for r in subset]
        st1_count = sum(1 for v in rank_vals
                        if str(v).strip() == "1")
        st1_rate = st1_count / len(subset) if subset else 0.0

        row = i + 3
        bg = BOAT_BG.get(boat_no, "AAAAAA")
        fg = BOAT_FG.get(boat_no, "000000")
        values = [
            f"{course}コース/{boat_no}号艇" if boat_no != course else str(course),
            name,
            len(subset),
            f"{avg_st:.3f}" if avg_st is not None else "—",
            f"{f_rate:.1%}",
            f"{st1_rate:.1%}",
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = center
            cell.border = border
            if ci == 1:
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.font = Font(bold=True, color=fg)
        ws.row_dimensions[row].height = 18

    # 行高さを明示してEMUを確定計算
    # 行1:28pt, 行2:20pt, 行3-8:18pt×6, 行9:6pt
    ROW_HEIGHTS_PT = [28, 20, 18, 18, 18, 18, 18, 18, 6]
    for r, h in enumerate(ROW_HEIGHTS_PT, 1):
        ws.row_dimensions[r].height = h
    bg_top_emu = int(sum(ROW_HEIGHTS_PT) * 12700)  # 1pt = 12700 EMU

    boat_info = _collect_boat_info(players_data)
    bg_png = _make_formation_background(boat_info, show_start_line=True)
    _embed_formation_in_sheet(ws, boat_info, bg_png, show_labels=True,
                               bg_top_emu=bg_top_emu)

    ws.column_dimensions["A"].width = 10


def _write_clean_formation_sheet(wb, players_data: list[dict],
                                   manual_offsets: dict | None = None):
    ws = wb.create_sheet("スタート配置図")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6
    bg_top_emu = int((28 + 6) * 12700)

    boat_info_raw = _collect_boat_info(players_data)
    boat_info = _apply_boat_corrections(boat_info_raw, manual_offsets)
    bg_png = _make_formation_background(boat_info, show_start_line=False)
    _embed_formation_in_sheet(ws, boat_info, bg_png, show_labels=False,
                               bg_top_emu=bg_top_emu)

    ws.column_dimensions["A"].width = 10


def _write_performance_sheet(wb, players_data: list[dict]):
    ws = wb.create_sheet("船足評価")
    ws.sheet_view.showGridLines = False

    png_data = make_boat_performance_chart(players_data)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(png_data)
        tmp_path = tmp.name
    img = XLImage(tmp_path)
    img.width  = 1000
    img.height = 420
    ws.add_image(img, "A1")
    ws.column_dimensions["A"].width = 130

